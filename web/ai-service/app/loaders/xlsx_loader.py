"""XLSX document loader.

Uses openpyxl to extract data from all sheets in an Excel workbook.
Opens workbooks in read-only mode (`read_only=True`, `data_only=True`) to
avoid evaluating formulas and to keep memory overhead low.

Each sheet is serialised as newline-separated rows in the content string.
Sheet metadata (name, row count, column names) is preserved in metadata.
"""
import io
import zipfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.loaders.base import BaseLoader
from app.loaders.exceptions import CorruptedFileError, EmptyFileError
from app.models.document import Document, FileType
from app.utils.metadata import clean_metadata


class XlsxLoader(BaseLoader):
    @property
    def file_type(self) -> str:
        return "xlsx"

    def load(self, content: bytes, filename: str) -> Document:
        if not content:
            raise EmptyFileError(f"XLSX file '{filename}' is empty.")

        try:
            wb = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
            raise CorruptedFileError(
                f"Cannot read XLSX '{filename}': file is corrupted or not a valid .xlsx."
            ) from exc
        except Exception as exc:
            raise CorruptedFileError(
                f"Unexpected error reading XLSX '{filename}': {exc}"
            ) from exc

        text_parts: list[str] = []
        sheet_meta: dict[str, Any] = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([str(cell) if cell is not None else "" for cell in row])

            if not all_rows:
                sheet_meta[sheet_name] = {"row_count": 0, "column_names": []}
                continue

            headers = all_rows[0]
            data_rows = all_rows[1:]

            sheet_meta[sheet_name] = {
                "row_count": len(data_rows),
                "column_names": headers,
            }

            text_parts.append(f"[Sheet: {sheet_name}]")
            text_parts.extend(
                "\t".join(row) for row in all_rows if any(cell.strip() for cell in row)
            )

        wb.close()

        content_str = "\n".join(text_parts)
        if not content_str.strip():
            raise EmptyFileError(
                f"XLSX file '{filename}' contains no readable data."
            )

        metadata = clean_metadata(
            {
                "sheet_names": list(wb.sheetnames),
                "sheet_count": len(wb.sheetnames),
                "sheets": sheet_meta,
            }
        )

        return Document(
            filename=filename,
            file_type=FileType.XLSX,
            mime_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            content=content_str,
            metadata=metadata,
            page_count=None,
            source=filename,
            size_bytes=len(content),
        )
